import base64
import csv
import io
import json
import re
import urllib.request
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from app.core.config import settings


class HistoryParseError(ValueError):
    pass


@dataclass
class ParsedCourse:
    course_code: str | None
    course_name: str
    semester: str | None = None
    credits: float | None = None


HEADER_ALIASES = {
    "course_code": {"课程号", "课程代码", "课号", "coursecode", "code"},
    "course_name": {"课程名称", "课程名", "名称", "coursename", "name"},
    "semester": {"学期", "修读学期", "semester", "term"},
    "credits": {"学分", "credits", "credit"},
}


def _normalise_header(value) -> str:
    return re.sub(r"[\s_\-:/]", "", str(value or "").strip().lower())


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HistoryParseError("Unable to decode the uploaded text file")


def _to_float(value) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group()) if match else None


def _rows_to_courses(rows: list[list]) -> list[ParsedCourse]:
    if not rows:
        return []
    header_index = None
    mapping: dict[str, int] = {}
    for index, row in enumerate(rows[:10]):
        candidate = {}
        for column, value in enumerate(row):
            normalised = _normalise_header(value)
            for field, aliases in HEADER_ALIASES.items():
                if normalised in aliases:
                    candidate[field] = column
        if "course_name" in candidate:
            header_index = index
            mapping = candidate
            break
    if header_index is None:
        return _parse_plain_lines([" ".join(str(cell or "") for cell in row) for row in rows])

    result = []
    for row in rows[header_index + 1 :]:
        name_index = mapping["course_name"]
        if name_index >= len(row):
            continue
        course_name = str(row[name_index] or "").strip()
        if not course_name:
            continue

        def get(field: str):
            column = mapping.get(field)
            return row[column] if column is not None and column < len(row) else None

        result.append(
            ParsedCourse(
                course_code=str(get("course_code") or "").strip() or None,
                course_name=course_name,
                semester=str(get("semester") or "").strip() or None,
                credits=_to_float(get("credits")),
            )
        )
    return result


def _parse_plain_lines(lines: list[str]) -> list[ParsedCourse]:
    records = []
    ignored = {"课程表", "历史课表", "课程名称", "课程名", "课程号", "学分", "教师", "上课时间"}
    for line in lines:
        cleaned = re.sub(r"\s+", " ", line.replace("|", " ").strip(" |-\t"))
        if not cleaned or cleaned in ignored or set(cleaned) <= {"-", ":"}:
            continue
        code_match = re.search(r"\b(?=[A-Za-z0-9]*\d)[A-Za-z0-9]{6,20}\b", cleaned)
        code = code_match.group(0) if code_match else None
        parts = [part.strip() for part in re.split(r"\t|\s{2,}|,", line) if part.strip()]
        name = None
        for part in parts:
            if code and part == code:
                continue
            if re.search(r"[\u4e00-\u9fff]", part) and not any(label in part for label in ignored):
                name = re.sub(r"\b\d+(?:\.\d+)?\s*学分\b", "", part).strip(" |-，,")
                break
        if not name and code:
            tail = cleaned.replace(code, "", 1).strip(" |-，,")
            name = re.split(r"\s+\d+(?:\.\d+)?(?:\s*学分)?\b", tail)[0].strip()
        if name and len(name) >= 2:
            records.append(
                ParsedCourse(
                    course_code=code,
                    course_name=name,
                    credits=_to_float(cleaned.split("学分")[0].split()[-1]) if "学分" in cleaned else None,
                )
            )
    return records


def _parse_csv(content: bytes) -> list[ParsedCourse]:
    text = _decode_text(content)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    return _rows_to_courses(rows)


def _parse_json(content: bytes) -> list[ParsedCourse]:
    data = json.loads(_decode_text(content))
    rows = data.get("courses", data) if isinstance(data, dict) else data
    if not isinstance(rows, list):
        raise HistoryParseError("JSON must contain a list of courses")
    result = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        name = item.get("course_name") or item.get("课程名称") or item.get("name")
        if not name:
            continue
        result.append(
            ParsedCourse(
                course_code=item.get("course_code") or item.get("课程号") or item.get("code"),
                course_name=str(name).strip(),
                semester=item.get("semester") or item.get("学期"),
                credits=_to_float(item.get("credits") or item.get("学分")),
            )
        )
    return result


def _parse_xlsx(content: bytes) -> list[ParsedCourse]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return _rows_to_courses(rows)


def _parse_pdf(content: bytes) -> list[ParsedCourse]:
    reader = PdfReader(io.BytesIO(content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not text.strip():
        raise HistoryParseError("PDF contains no extractable text; configure OCR or upload an image")
    lines = text.splitlines()
    table_like = [re.split(r"\s{2,}|\t", line.strip()) for line in lines if line.strip()]
    records = _rows_to_courses(table_like)
    return records or _parse_plain_lines(lines)


def _parse_image(content: bytes, suffix: str) -> list[ParsedCourse]:
    if not settings.HISTORY_OCR_ENDPOINT:
        raise HistoryParseError(
            "Image OCR is not configured; configure HISTORY_OCR_ENDPOINT or use manual entry"
        )
    request_body = json.dumps(
        {
            "mime_type": f"image/{suffix.lstrip('.').replace('jpg', 'jpeg')}",
            "content_base64": base64.b64encode(content).decode("ascii"),
        }
    ).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if settings.HISTORY_OCR_API_KEY:
        headers["Authorization"] = f"Bearer {settings.HISTORY_OCR_API_KEY}"
    request = urllib.request.Request(
        settings.HISTORY_OCR_ENDPOINT, data=request_body, headers=headers, method="POST"
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            result = json.loads(response.read().decode("utf-8"))
    except Exception as exc:
        raise HistoryParseError(f"OCR request failed: {exc}") from exc
    if isinstance(result.get("records"), list):
        return _parse_json(json.dumps(result["records"], ensure_ascii=False).encode("utf-8"))
    text = result.get("text", "")
    return _parse_plain_lines(text.splitlines())


def parse_history_file(file_name: str, content: bytes) -> tuple[str, list[ParsedCourse], list[str]]:
    suffix = Path(file_name).suffix.lower()
    warnings: list[str] = []
    if suffix in {".csv", ".tsv"}:
        parser_name, records = "tabular_csv_v1", _parse_csv(content)
    elif suffix == ".json":
        parser_name, records = "json_v1", _parse_json(content)
    elif suffix in {".xlsx", ".xlsm"}:
        parser_name, records = "xlsx_v1", _parse_xlsx(content)
    elif suffix == ".pdf":
        parser_name, records = "pdf_text_v1", _parse_pdf(content)
    elif suffix in {".txt", ".md"}:
        parser_name, records = "plain_text_v1", _parse_plain_lines(_decode_text(content).splitlines())
    elif suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        parser_name, records = "ocr_adapter_v1", _parse_image(content, suffix)
        warnings.append("OCR results may be inaccurate and require user confirmation")
    else:
        raise HistoryParseError(f"Unsupported file type: {suffix or 'unknown'}")
    if not records:
        raise HistoryParseError("No course records could be parsed from the uploaded file")
    return parser_name, records, warnings
